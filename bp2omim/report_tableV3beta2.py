#!/usr/bin/python
#coding=UTF-8

#report table subscript
##############################################################
###Dependencies:
import sys
sys.path.append('/home/analise_liWGSseqs/software/biomart')
sys.path.append("/usr/lib/pymodules/python2.7/openpyxl")
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from biomart import BiomartServer
import urllib2
from urllib2 import Request, urlopen, URLError
from collections import OrderedDict
import time
import random
import get_categories


def parse_HI(infile):
        """parses the HI file and retrives a dictionary with
        dic[gene]=HI"""
        f=open(infile)
        f.readline()
        dic={}
        for i in f:
                line=i.split("\t")
                camp=line[3].split("|")
                dic[camp[0]]=camp[-1][:-1]
        f.close()
        return dic

def select_tads(tads, chrr, bp, chrsz):
	"""using the TADs file, reference chromossome and breakpoint, searchs for the breakpoint TAD on the list.
	In case of the breakpoint beeing located between TADs, retrives the interval that reunites the flanking TADs.
	Returns a dictionary as dic[TAD]=[int(start), int(end)]"""
	f=open(tads)
	tadr=[]
	tad3={}
	ind=-1
	a=0
	sz=read_chr_size(chrsz)
	for i in f:
		line=i.split()
		if line[0]=="chr"+chrr:#alt
			tadr.append(line[1]+" "+line[2])
			if int(bp)>= int(line[1]) and int(bp)<= int(line[2]):
				ind=tadr.index(line[1]+" "+line[2])
	if int(bp)<int(tadr[0].split(" ")[0]):
		tad3["pter_br"]=["1", tadr[0].split(" ")[0]]
		tad3["TAD+1"]=[tadr[1].split(" ")[0], tadr[1].split(" ")[1]]
		return tad3
	if int(bp)>int(tadr[-1].split(" ")[1]):
		tad3["TAD-1"]=[tadr[-1].split(" ")[0], tadr[-1].split(" ")[1]]
		tad3["qter_br"]=[tadr[-1].split(" ")[1], sz[chrr]]
		return tad3
	if ind==-1:
		temp=""
		for el in tadr:
			if temp=="":
				temp=int(el.split()[1])
			else:
				if int(bp)>temp and int(bp)< int(el.split()[0]):
					ind=tadr.index(el)
					a=ind-1
					break
				else:
					temp=int(el.split()[1])
	if a==0:#so for entre TADs da a coordenada inicial da tad anterior e a coordenada de final da tad posterior
		if ind!=0:
			tad3["TAD-1"]=[(tadr[ind-1].split(" ")[0])]#tad-1 start
			tad3["TAD-1"].append(tadr[ind-1].split(" ")[1])#tad-1 end
		try:
			tad3["brTAD"]=[(tadr[ind].split(" ")[0])]#brTAD
			tad3["brTAD"].append(tadr[ind].split(" ")[1])#brTAD
		except IndexError:
			pass
		try:
			tad3["TAD+1"]=[(tadr[ind+1].split(" ")[0])]#tad+1
			tad3["TAD+1"].append(tadr[ind+1].split(" ")[1])#tad+1
		except IndexError:
			pass
	else:
		try:
			tad3["TAD-1"]=[(tadr[a].split(" ")[0])]
			tad3["TAD-1"].append(tadr[a].split(" ")[1])
		except IndexError:
			pass
		tad3["TAD+1"]=[(tadr[ind].split(" ")[0])]
		tad3["TAD+1"].append(tadr[ind].split(" ")[1])
	f.close()
	return tad3

def parse_tads(tad3):
	""" parses the tad information from select tads, and parses the interest tads
	to a new dictionary with the same structure dic[TAD]=[int(start), int(end)]"""
	if "TAD-1" in tad3 and "TAD+1" in tad3 and "brTAD" in tad3:
		if int(tad3["brTAD"][0])-int(tad3["TAD-1"][1])>5:
			tad3["TBR_1"]=[tad3["TAD-1"][1],tad3["brTAD"][0]]
		if int(tad3["TAD+1"][0])-int(tad3["brTAD"][1])>5:
			tad3["TBR_2"]=[tad3["brTAD"][1],tad3["TAD+1"][0]]
	elif "TAD-1" in tad3 and "TAD+1" in tad3 and "brTAD" not in tad3:
		if int(tad3["TAD+1"][0])-int(tad3["TAD-1"][1])>5:
			tad3["TBR"]=[tad3["TAD-1"][1],tad3["TAD+1"][0]]
	elif "TAD-1" not in tad3 and "TAD+1" in tad3 and "brTAD" in tad3:
		if int(tad3["TAD+1"][0])-int(tad3["brTAD"][1])>5:
			tad3["TBR_2"]=[tad3["brTAD"][1],tad3["TAD+1"][0]]
	elif "TAD-1" in tad3 and "TAD+1" not in tad3 and "brTAD" in tad3:
		if int(tad3["brTAD"][0])-int(tad3["TAD-1"][1])>5:
			tad3["TBR_1"]=[tad3["TAD-1"][1],tad3["brTAD"][0]]
	new={}
	if "pter_br" in tad3:
		new["pter_br_"]=tad3["pter_br"]
		tdds.append("pter_br_")
	if "TAD-1_" in tdds and "TAD-1" in tad3:
		new["TAD-1_"]=tad3["TAD-1"]
	if "TAD+1_" in tdds and "TAD+1" in tad3:
		new["TAD+1_"]=tad3["TAD+1"]
	if "TBR" in tad3:
		new["TBR_"]=tad3["TBR"]
		tdds.append("TBR_")############################################
	if "TBR" not in tad3 and "brTAD" in tad3:
		new["brTAD_"]=tad3["brTAD"]
		tdds.append("brTAD_")############################################
	if "qter_br" in tad3:
		new["qter_br_"]=tad3["qter_br"]
		tdds.append("qter_br_")	
	return new

def search(filt, att, version):
	"""makes the search against the biomart database, using as filter arguments
	filt and retriving the atributes at att"""
	if version=="hg19":
		server = BiomartServer( "http://feb2014.archive.ensembl.org/biomart/" )
		hg=server.datasets["hsapiens_gene_ensembl"]
	elif version=="hg38":
		server=BiomartServer( "http://jul2018.archive.ensembl.org/biomart/" )
		hg=server.datasets["hsapiens_gene_ensembl"]
	response = hg.search({
	  'filters':filt,
	  'attributes': att
	})
	return response
		
		
def parse_first_search(response, filt, bp, chrr,key,numb):
	"""Parses the output from search, and returns two dictionaries
	one with the genomic elements disrupted by the breakpoint, as
	disrupted[gene/lincRNA..]=[chr:start-end, description/lincRNA_Ensembl_entry_name, strand, gene biotype, ensemblID]
	and another with all the genomic elements
	dic[gene]=[ch:start-end, description, strand, NOOMIM/OMIM ID, gene biotype, ensembl ID]"""
	#	at1=[ 'start_position', 'end_position','external_gene_id', 'strand', 'mim_gene_accession', 'gene_biotype', 'ensembl_gene_id']
	dic={}
	disrupted={}
	aa=set()
	for line in response.iter_lines():
		d=line.split("\t")
		s=int(d[0])
		e=int(d[1])
		if int(d[0])<int(bp) and int(d[1])>int(bp):
			if d[-2]=="lincRNA":
				disrupted[d[2].split(".")[0]]=["chr"+chrr+":"+'{:0,}'.format(s)+"-"+'{:0,}'.format(e), d[3], d[5], d[6].strip()]#alt
			else:
				disrupted[d[2].split(".")[0]]=["chr"+chrr+":"+'{:0,}'.format(s)+"-"+'{:0,}'.format(e), d[3], d[5], d[6].strip()]#alt
		if len(d[-3])>2:
			if d[2].split(".")[0] not in aa:
				dic[int(d[0])]=[d[2].split(".")[0],"chr"+chrr+":"+'{:0,}'.format(s)+"-"+'{:0,}'.format(e), d[3], d[-3].strip(), d[5], d[6].strip()]#alt
				aa.add(d[2].split(".")[0])
		elif len(d[-3])<2:
			if d[2].split(".")[0] not in aa:
				if d[-2]=="lincRNA":
					dic[int(d[0])]=[d[2].split(".")[0],"chr"+chrr+":"+'{:0,}'.format(s)+"-"+'{:0,}'.format(e), d[3], "NOOMIM", d[5], d[6].strip()]#alt
				else:
					dic[int(d[0])]=[d[2].split(".")[0],"chr"+chrr+":"+'{:0,}'.format(s)+"-"+'{:0,}'.format(e), d[3], "NOOMIM", d[5], d[6].strip()]#alt
				aa.add(d[2].split(".")[0])
	if int(filt["start"][0])<=int(bp) and int(filt["end"][0])>=int(bp):
		bb=int(bp)
		dic[int(bp)]=["Breakpoint"+numb, "Chr"+chrr+":"+'{:0,}'.format(bb), "", "","NOOMIM", "", ""]
	od = OrderedDict(sorted(dic.items()))
	final=OrderedDict()
	for key, value in od.items():
		final[value[0]]=value[1:]
	return final,disrupted


def get_omim(dicc, disrupted,bp, tadd, dic_text, chrrr, name):
	"""uses the information from parse_first_search, the breakpoint, TAD, and disrupted genes information (dic_text)
	and makes two steps:
	1 - the genomic elements without OMIM are sended to make_report_table_not_OMIM, where the information is organized to the final output
	2 - the genomic elements with OMIM are search on OMIM database, using the OMIM API, and sended to make_report_table, to organize the
	information and write the output"""
	for key, value in dicc.items():
		if value[-3]=="NOOMIM" and tadd in tdds:
			make_report_table_not_OMIM(key,dicc, dic_text, chrrr, name)
		if value[-2]!="BREAKPOINT" and value[-3]!="NOOMIM":###############################
			request = Request("http://api.omim.org/api/entry?mimNumber="+value[-3]+"&apiKey="+"API_KEY"+"&format=python&include=geneMap")
			try:
				response = urlopen(request)
				kittens = response.read()
				a=kittens.split("\n")
				c=replace_series(replace_series(replace_series(replace_series(replace_series(a,"{"), "}"),"'"),"["),"]")
				dic={}
				for el in c:
					if el.startswith("mimNumber"):
						ss=(el.split(":")[-1]).strip()
						dic["mimNumber"]=['=HYPERLINK("http://omim.org/entry/'+ss.strip(",")+'","'+ss.strip(",")+'")']
					elif el.startswith("mouseMgiID"):
						ss=(el.split(":")[-1]).strip()
						dic["mouseMgiID"]='=HYPERLINK("http://www.informatics.jax.org/allele/summary?markerId=MGI:'+ss.strip(",")+'")'
					elif el.startswith("geneName"):
						ss=(el.split(":")[-1]).strip()
						dic["geneName"]=ss.strip(",")
					elif el.startswith("phenotypeMimNumber"):
						ss=(el.split(":")[-1]).strip()
						if "phenotypeMimNumber" not in dic:
							dic["phenotypeMimNumber"]=['=HYPERLINK("http://omim.org/entry/'+ss.strip(",")+'","'+ss.strip(",")+'")']
						else:
							dic["phenotypeMimNumber"].append('=HYPERLINK("http://omim.org/entry/'+ss.strip(",")+'","'+ss.strip(",")+'")')
					elif el.startswith("phenotypeInheritance"):
						inn=(el.split(":")[-1])
						if "phenotypeInheritance" not in dic:
							dic["phenotypeInheritance"]=[inn]
						else:
							dic["phenotypeInheritance"].append(inn)
					elif el.startswith("phenotype:"):
						ss=(el.split(":")[-1]).strip()
						if "phenotype" not in dic:
							dic["phenotype"]=[ss.strip(",")]
						else:
							if "phenotypeMimNumber" not in dic:
								dic["phenotypeMimNumber"]=["na"]
							if "phenotypeInheritance" not in dic:
								dic["phenotypeInheritance"]=["nd"]
							if len(dic["phenotype"])!=len(dic["phenotypeMimNumber"]):
								dic["phenotypeMimNumber"].append("na")
							if len(dic["phenotype"])!=len(dic["phenotypeInheritance"]):
								dic["phenotypeInheritance"].append("nd")
							dic["phenotype"].append(ss.strip(","))
				if "geneName" not in dic:
					dic["geneName"]="NA"
				if "mouseMgiID" not in dic:
					dic["mouseMgiID"]=""
				if "phenotype" in dic:
					if "phenotypeMimNumber" not in dic:
						dic["phenotypeMimNumber"]=["na"]
					if len(dic["phenotype"])!=len(dic["phenotypeMimNumber"]):
						dic["phenotypeMimNumber"].append("na")
				if tadd in tdds:
					make_report_table(key,dic, dic_text, dicc, chrrr, name)
			except URLError, e:
				print 'No kittez. Got an error code:', e

def replace_series(l, what):
	"""small method used in get_omim to replace what by nothing"""
	aa=[]
	for e in l:
		aa.append(e.replace(what,""))
	return aa
				
def make_report_table(name, dic, dic_text, dicc, chrrr, name2):
	HI=parse_HI("HI_Predictions_Version3.bed")
	dd2p=get_categories.get_dd2p("DDG2P_21_2_2019.csv")
	"""Uses the information from get_omim to organize the output, and write it.
	USED ONLY IN THE GENOMIC ELEMENTS WITH OMIM ID ASSOCIATED"""
	dd2pp=[]
	if name in dd2p:
		dd2pp=dd2p[name]
	if name in HI:
		hplo=HI[name]
	if name not in HI:
		hplo="nd"
	if name in dic_text:
		pli=parse_oe("oe", name, True)
		aa="http://www.genecards.org/cgi-bin/carddisp.pl?gene="+dicc[name][-1]
		if check_genecard_existence(aa)=="true":
			if len(dic_text[name])<=3:
				genecard=['=HYPERLINK("http://www.genecards.org/cgi-bin/carddisp.pl?gene='+dicc[name][-1]+'", "'+name+' '+dic_text[name][0]+'")']
			else:
				genecard=['=HYPERLINK("http://www.genecards.org/cgi-bin/carddisp.pl?gene='+dicc[name][-1]+'", "'+name+' '+dic_text[name][0]+'")']
		else:
			if len(dic_text[name])<=3:
				genecard=[name+' '+dic_text[name][0]]
			else:
				genecard=[name+' '+dic_text[name][0]]
	if name not in dic_text:
		pli=parse_oe("oe", name, False)
		aa="http://www.genecards.org/cgi-bin/carddisp.pl?gene="+dicc[name][-1]
		if check_genecard_existence(aa)=="true":
			genecard=['=HYPERLINK("http://www.genecards.org/cgi-bin/carddisp.pl?gene='+dicc[name][-1]+'", "'+name+'")']
		else:
			genecard=[name]
	if "phenotypeMimNumber" in dic:#######################adicionar uma maneira de ele ler as refs
		count=0
		while (count<len(genecard)):
			aaa=0
			while aaa<len(dic["phenotypeInheritance"]):
				inh=dic["phenotypeInheritance"][aaa]
				if dic["phenotypeInheritance"][aaa]==" None":
					inh="nd"
				if dic["phenotypeInheritance"][aaa]==" Autosomal recessive":
					inh="AR"
				if dic["phenotypeInheritance"][aaa]==" Autosomal dominant":
					inh="AD"
				if aaa==0:
					ws2.append([genecard[count],dic["mimNumber"][0],pli, hplo, ' ', dic['phenotype'][aaa], dic['phenotypeMimNumber'][aaa], inh])
					aaa+=1
				else:
					ws2.append(["","","","","",dic['phenotype'][aaa], dic['phenotypeMimNumber'][aaa],inh])
					aaa+=1
			if count==0 and name in dic_text:
				ws2.append([name2])
			count+=1
	else:
		if "mimNumber" in dic:
			ws2.append([genecard[0],dic["mimNumber"][0], pli,hplo,"","-","na","nd"])
			if name in dic_text:
				ws2.append([name2])

def parse_pli(pli):
	f=open(pli)
	f.readline()
	pli_dic={}
	for el in f:
		line=el.split("\t")
		pli_dic[line[1]]="{0:.2f}".format(round(float(line[-1]),2))
	f.close()
	return pli_dic

def parse_oe(oe, ell, is_dis):
	f=open(oe)
	f.readline()
	nn="nd"
	for el in f:
		line=el.split("\t")
		if line[0]==ell:
			if is_dis==True or float(line[-1].split("(")[0])<=0.35:########
				nn=line[-1].strip()
			else:
				nn=line[-1].split("(")[0]
	f.close()
	return nn

def make_report_table_not_OMIM(el,dic, dic_text, chrrr, name):
	HI=parse_HI("HI_Predictions_Version3.bed")
	
	"""called on get_omim to organize the output, and write it.
	USED ONLY IN THE GENOMIC ELEMENTS WITHOUT OMIM ID ASSOCIATED"""
	dicd=dic_text.keys()
	if el in HI:
		hplo=HI[el]
	if el not in HI:
		hplo="nd"
	if el.startswith("Breakpoint") and len(dic_text)==0:
		ws2.append([name])
	if len(dic_text)>0:
		if el.startswith("Breakpoint_A") and "Breakpoint_B" in dic.keys():
			if abs(dic.keys().index("Breakpoint_A")-dic.keys().index(dicd[0]))>len(dicd):
				ws2.append([name])
		if el.startswith("Breakpoint_B") and "Breakpoint_A" in dic.keys():
			if abs(dic.keys().index("Breakpoint_B")-dic.keys().index(dicd[0]))>len(dicd):
				ws2.append([name])	
	if el in dic_text:
		aa="http://www.genecards.org/cgi-bin/carddisp.pl?gene="+dic[el][-1]
		if check_genecard_existence(aa)=="true":
			genecard='=HYPERLINK("http://www.genecards.org/cgi-bin/carddisp.pl?gene='+dic[el][-1]+'", "'+el+' '+dic_text[el][0]+'")'
		else:
			genecard=name+' '+dic_text[name][0]
		pli=parse_oe("oe", el, True)
		ws2.append([genecard,"na",pli,hplo,"", "-","na", "nd"])
		ws2.append([name])
	elif len(el.split(" "))>1:
		gg=el.split(" ")[1]
		if gg in dic_text:
			aa="http://www.genecards.org/cgi-bin/carddisp.pl?gene="+dic[el][-1]
			if check_genecard_existence(aa)=="true":
				genecard='=HYPERLINK("http://www.genecards.org/cgi-bin/carddisp.pl?gene='+dic[el][-1]+'", "'+el+' '+dic_text[gg][0]+'")'
			else:
				genecard=el+' '+dic_text[gg][0]
			pli=parse_oe("oe", el, True)
			ws2.append([genecard, "na",pli,hplo,"", "-","na", "nd"])
			ws2.append([name])
	elif el not in dic_text and (dic[el][-2]=="protein_coding" or dic[el][-2]=="lincRNA"):
		aa="http://www.genecards.org/cgi-bin/carddisp.pl?gene="+dic[el][-1]
		if check_genecard_existence(aa)=="true":
			genecard='=HYPERLINK("http://www.genecards.org/cgi-bin/carddisp.pl?gene='+dic[el][-1]+'", "'+el+'")'
		else:
			genecard=el
		pli=parse_oe("oe", el, False)
		ws2.append([genecard, "na",pli,hplo,"", "-","na", "nd"])

def interr(filt1,version):
	"""uses the search method to search for the disrupted genomic elements and returns a dictionary with
	the exons, as dic[ensemble_transcript_id]=[exonstart,exonend,exonstart,exonend...,transcript_name, gene_name, strand]"""
	dic={}
	if version=="hg19":
		at2=['ensembl_transcript_id', 'exon_chrom_start','exon_chrom_end', 'rank']
		at3=['external_transcript_id','external_gene_id','ensembl_transcript_id', "strand"]
	elif version=="hg38":
		at2=['ensembl_transcript_id', 'exon_chrom_start','exon_chrom_end', 'rank']
		at3=['external_transcript_name','external_gene_name','ensembl_transcript_id', "strand"]
	fstep=search(filt1,at2,version)
	for line in fstep.iter_lines():
		i=line.split()
		if i[0] not in dic:
			dic[i[0]]=[i[1],i[2]]
		else:
			dic[i[0]].append(i[1])
			dic[i[0]].append(i[2])
	for key,value in dic.items():
		value.sort()
	sstep=search({'link_ensembl_transcript_stable_id':dic.keys()}, at3,version)
	for line in sstep.iter_lines():#poe num dic a parte e quando escreve, se tiver no dic ele adiciona info
		i=line.split()
		if i[-2] in dic:
			dic[i[-2]].append(i[0])
			dic[i[-2]].append(i[1])
                        dic[i[-2]].append(i[-1])
	return dic#dic[ensemble_transcript_id]=[exonstart,exonend,exonstart,exonend...,transcript_name, gene_name, strand]

def ivsreport(dic_pos, bp, bp2):
	"""Uses the dictionary from interr and the breakpoint to retrive a dictionary with the specific position of the
	breakpoint in the disrupted gene. Returns a dictionary as dic[gene name]=[Exon 1 - IVS7, IVS7 - Exon 9]"""
	dic_text={}
	for key, value in dic_pos.items():
		if int(value[-4])>=int(bp) and int(value[0])<=int(bp):
			el=0
			while el<len(value[:-4]):
				if int(bp)>=int(value[el]) and int(bp)<int(value[el+1]):
					if bp2=="" or int(bp2)<int(value[el]) or int(bp2)>int(value[el+1]):#########################
						if el%2==1:#impar logo intrao
							if value[-1]=="1":
								intrao_numb=el/2+1
								aa=[" - IVS"+str(intrao_numb), bp]
							else:
								intrao_numb=(len(value[:-3])/2)-(el/2+1)
								aa=[" - IVS"+str(intrao_numb), bp]
						else:
							if value[-1]=="1":
								exon_numb=el/2+1
								aa=["- Exon "+str(exon_numb), bp]
							else:
								exon_numb=(len(value[:-3])/2)-(el/2+1)
								aa=[" - Exon "+str(exon_numb), bp]
					else:
						excount=len(value[:-4])/2
						if el%2==1: #intrao
							if value[-1]=="1":
								intrao_numb=el/2+1
								aa=[" - IVS"+str(intrao_numb), bp]
							else:
								intrao_numb=(len(value[:-3])/2)-(el/2+1)
								aa=[" - IVS"+str(intrao_numb),bp]
						else:
							if value[-1]=="1":
								exon_numb=el/2+1
								aa=[" - Exon "+str(exon_numb), bp]
							else:
								exon_numb=(len(value[:-3])/2)-(el/2+1)
								aa=[" - Exon "+str(exon_numb),bp]
					dic_text[value[-2].split(".")[0]]=aa
				el+=1
	return dic_text
	
def final_things_sub(ch, value, bp, key, version, numb, bp2):
	interromped={}
	filt1= {'chromosome_name': [ch], 'start': [str(value[0])], 'end':[str(value[1])]}#alt
	dic,disrupted=parse_first_search(search(filt1, at1,version),filt1,bp,ch,key, numb)#alt
	dic_text={}
	if len(disrupted)!=0:
		dic_text=ivsreport(interr({'chromosome_name': [ch], 'start': [bp], 'end':[str(int(bp)+1)]},version), bp, bp2)
	return dic, disrupted, dic_text
	
	
def final_things(ts, ch, bp,version, twobps, is_region, bb, name):#16q24.3 breakpoint within the hESC TAD at chr16:89246091-89686091
	"""Used by the deal_with... methods to run the whole search of the genomic elements"""
	for key, value in ts.items():
		if (key=="brTAD_" or key=="TBR_") and twobps==True:
			dic1, disrupted1, dic_text1=final_things_sub(ch, value, bp[0], key, version,"_A", bp[1])
			dic2, disrupted2, dic_text2=final_things_sub(ch, value, bp[1], key, version, "_B", "")
			dic=merge_two_ordered_dicts(dic1, dic2)
			disrupted=merge_two_dicts(disrupted1, disrupted2)
			dic_text=merge_two_dicts(dic_text1, dic_text2)
		elif twobps==False:
			dic, disrupted, dic_text=final_things_sub(ch, value, bp, key, version, "", "")
		elif key!="brTAD_" and key!="TBR_":
			dic, disrupted, dic_text=final_things_sub(ch, value, bp[0], key, version, "", "")
		if key in tdds:
			if key=="pter_br_":
				cor="pter region at "+ch+":"+value[0]+"-"+value[1]
			if key=="qter_br_":
				cor="qter region at "+ch+":"+value[0]+"-"+value[1]
			if key=="brTAD_":
				cor=ch+bb+" breakpoint within the "+tad+"TAD at "+ch+":"+value[0]+"-"+value[1]
			elif key=="TAD-1_" or key=="TAD+1_":
				cor=tad+" "+key[:-1]+" at chr"+ch+":"+value[0]+"-"+value[1]
			elif key=="TBR_":
				cor=ch+bb+" breakpoint within the "+tad+" interTAD region at chr"+ch+":"+value[0]+"-"+value[1]
			elif key.startswith("TDel"):
				cor=tad+" TAD at chr"+ch+":"+str(value[0])+"-"+str(value[1])
			elif key.startswith("interTDel"):
				cor=tad+" interTAD region at chr"+ch+":"+str(value[0])+"-"+str(value[1])
			ws2.append([cor, "", "","",""])
		get_omim(dic, disrupted, bp, key, dic_text, ch, name)

def merge_two_dicts(x, y):
    """Given two dicts, merge them into a new dict as a shallow copy."""
    z = y.copy()
    z.update(x)
    return z

def merge_two_ordered_dicts(x,y):
	z=OrderedDict()
	kh=y.keys()
	bpp=kh.index("Breakpoint_B")
	aa=0
	for key, value in x.items():
		if aa!=bpp+1:
			z[key]=value
			aa+=1
		else:
			z["Breakpoint_B"]=y["Breakpoint_B"]
			z[key]=value
			aa+=1
	return z
	

def ordain(dicd):
	"""order the TADS by logic. used in deals_with_deletions and deal_with_translocations_inversions"""
	aa=OrderedDict()
	if "pter_br_" in dicd:
		aa["pter_br_"]=dicd["pter_br_"]
	if "TAD-1_" in dicd:
		aa["TAD-1_"]=dicd["TAD-1_"]
	if "brTAD_" in dicd:
		aa["brTAD_"]=dicd["brTAD_"]
	if "TBR_" in dicd:
		aa["TBR_"]=dicd["TBR_"]
	if "TAD+1_" in dicd:
		aa["TAD+1_"]=dicd["TAD+1_"]
	if "qter_br_" in dicd:
		aa["qter_br_"]=dicd["qter_br_"]
	return aa


def get_tdel(intnewt1, intnewt2):
	"""returns the delected region, when it is more than a TAD. used by deals_with_deletions"""
	if "pter_br_" in intnewt1:
		co1=intnewt1["pter_br_"][1]
	if "qter_br_" in intnewt2:
		co2=intnewt2["qter_br_"][0]
	if "TBR_" in intnewt1:
		co1=intnewt1["TBR_"][1]
	if "brTAD_" in intnewt1: 
		co1=intnewt1["brTAD_"][1]
	if "TBR_" in intnewt2:
		co2=intnewt2["TBR_"][0]
	if "brTAD_" in intnewt2: 
		co2=intnewt2["brTAD_"][0]
	if int(co2)-int(co1)>100:
		return {"TDel":[co1,co2]}
	else:
		return {}

def deleted_tads(tadd, tdel, ch):
	f=open(tadd)
	r=[]
	selected=OrderedDict()
	tads_inter=OrderedDict()
	numb=1
	for i in f:
		line=i.split("\t")
		if line[0]=="chr"+str(ch):
			if len(r)==0:
				r=[int(line[1]), int(line[2])]
				tads_inter["TDel_"+str(numb)]=[int(line[1]), int(line[2])]
				numb+=1
			else:
				if int(line[1])-r[1]>100:
					tads_inter["interTDel_"+str(numb)]=[int(r[1]), int(line[1])]
					numb+=1
					r=[int(line[1]), int(line[2])]
					tads_inter["TDel_"+str(numb)]=[int(line[1]), int(line[2])]
					numb+=1
				else:
					r=[int(line[1]), int(line[2])]
					tads_inter["TDel_"+str(numb)]=[int(line[1]), int(line[2])]
					numb+=1	
	f.close()
	for key, value in tads_inter.items():
		if int(tdel[0])<=value[0] and int(tdel[1])>=value[1]:
			selected[key]=value
			tdds.append(key)
	return selected
			

def deals_with_deletions(ch, br, version, tadd, is_region, b1, b2, name, chrsz):
	"""method used by get_results to deal with deletions"""
	newt1=parse_tads(select_tads(tadd, ch[0], int(br[0]),chrsz))#alt
	newt2=parse_tads(select_tads(tadd, ch[0], int(br[1]),chrsz))#alt
	intnewt1=ordain(newt1)
	intnewt2=ordain(newt2)
	if intnewt1!=intnewt2:
		tdel=get_tdel(intnewt1, intnewt2)###Não lida com o facto de poderem ser varias TADs, poe tudo no mesmo espaço
		if tdel!={}:
			final_things(intnewt1, ch[0], br[0], version, False, is_region, b1, name)
			newtds=deleted_tads(tadd, tdel["TDel"], ch[0])
			final_things(newtds, ch[0], "0",version, False, is_region, b1, name)#alt
			final_things(intnewt2,ch[0], br[1], version, False, is_region, b2, name)#alt
		else:
			if "TAD+1_" in intnewt1:
				del intnewt1["TAD+1_"]
			if "TAD-1_" in intnewt2:
				del intnewt2["TAD-1_"]
			final_things(intnewt1, ch[0], br[0], version, False, is_region, b1, name)	
			final_things(intnewt2,ch[0], br[1], version, False, is_region, b2, name)#alt
	else:
		final_things(intnewt1, ch[0], br, version, True, is_region, b1, name)

def deals_with_trans_with_dels(br, tadd, ch, version, newname1, newname2, b1, b2, name1, name2,chrsz):
	"""method used by get_results to deal with translocations/inversions associated with deletions"""
	br1=br[0].split("-")
	br2=br[1].split("-")
	if len(br1)==1:#se for só um valor, faz normalmente
		prepare_exel(newname1, True)
		newt1=parse_tads(select_tads(tadd, ch[0], int(br1[0]),chrsz))#alt
		intnewt1=ordain(newt1)	
		final_things(intnewt1, ch[0], br1[0],version, False, False, b1, name1)#alt
		if ch[0]!=ch[1]:
			make_format_ws2(ws2, False, False)
	if len(br1)>1:
		prepare_exel(newname1, True)
		chh=[ch[0]]
		deals_with_deletions(chh, br1, version, tadd, False, b1, b1, name1, chrsz)###
		if ch[0]!=ch[1]:
			make_format_ws2(ws2, True, False)
	if len(br2)>1:
		if newname2!="":
			prepare_exel(newname2, False)
		chh=[ch[1]]
		deals_with_deletions(chh, br2, version, tadd, False, b2, b2, name2, chrsz)###
		make_format_ws2(ws2, True, False)
	if len(br2)==1:#se for só um valor, faz normalmente
		if newname2!="":
			prepare_exel(newname2, False)
		newt2=parse_tads(select_tads(tadd, ch[1], int(br2[0]),chrsz))#alt
		intnewt2=ordain(newt2)	
		final_things(intnewt2, ch[1], br2[0],version, False, False, b2, name2)#alt
		make_format_ws2(ws2, False, False)


def read_chr_size(infile):
	f=open(infile)
	dic={}
	for el in f:
		s=el.split("\t")
		dic[s[0]]=s[1]
	f.close()
	return dic
		
def get_results(cA, cB, brA, brB, version, taddi, is_region, tt, der):
	name1=""
	name2=""
	"""the method that deals with the information from the formulae"""
	if version=="hg19":
		bandd="cytoband_hg19.bed"
		vista="vista_enhacers"
		chrsz="hg19.sizes"
	if version=="hg38":
		bandd="cytoband_hg38.bed"
		vista=""
		chrsz="hg38.sizes"
	ch=[cA,cB]
	br=[brA,brB]
	if tt=="Insertion":#chr0 donor chr1 recipent
		b1=read_cytoband(bandd, ch[0], br[0].split("-")[0])#alt
		b2=read_cytoband(bandd, ch[1], br[1].split("-")[0])#alt
		name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "ins", ch[0])
		name2=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "ins", ch[1])
		newname1=["Table 1 - Characterization of the insertion region "+ch[0]+b1,"","",""]
		newname2=["Table 2 - Characterization of the excision region "+ch[1]+b2,"","",""]
		deals_with_trans_with_dels(br, tadd, ch, version,newname1, newname2, b1, b2, name1, name2,chrsz)
	elif tt=="Deletion":#se o nome do segundo cromossoma for igual a "del", o programa assume que temos uma deleção entre a primeira e a segunda coordenada
		b1=read_cytoband(bandd, ch[0], br[0])#alt
		b2=read_cytoband(bandd, ch[0], br[1])#alt
		name1=pq_nomenclature(ch[0], ch[0], b1, b2, br[0], br[1], False, False, False, False, False, "del", "")
		newname=["Table 1 - Characterization of the deleted region "+"del("+ch[0]+")("+b1+";"+b2+")","","",""]
		prepare_exel(newname, True)
		deals_with_deletions(ch, br, version, tadd, is_region, b1,b2, name1,chrsz)#################
		make_format_ws2(ws2, True, is_region)
	elif tt=="Specific_region":#se o nome do segundo cromossoma for igual a "del", o programa assume que temos uma deleção entre a primeira e a segunda coordenada
		b1=read_cytoband(bandd, ch[0], br[0])#alt
		b2=read_cytoband(bandd, ch[0], br[1])#alt
		name=pq_nomenclature(ch[0], ch[0], b1, b2, br[0], br[1], True, False, False, False, False, "del", "")
		newname=["Table 1 - Characterization of the region of interest "+ch[0]+")("+b1+";"+b2+")","","",""]
		prepare_exel(newname, True)
		deals_with_deletions(ch, br, version, tadd, is_region, b1,b2, name, chrsz)#################
		make_format_ws2(ws2, True, is_region)
	elif tt=="Duplication":#se o nome do segundo cromossoma for igual a "del", o programa assume que temos uma deleção entre a primeira e a segunda coordenada
		b1=read_cytoband(bandd, ch[0], br[0])#alt
		b2=read_cytoband(bandd, ch[0], br[1])#alt
		name1=pq_nomenclature(ch[0], ch[0], b1, b2, br[0], br[1], False, False, False, False, False, "dup", "")
		newname=["Table 1 - Characterization of the duplicated region "+"dup("+ch[0]+")("+b1+";"+b2+")","","",""]
		prepare_exel(newname, True)
		deals_with_deletions(ch, br, version, tadd, is_region, b1,b2, name1, chrsz)#################
		make_format_ws2(ws2, True, is_region)
	else:
		if len(br[0].split("-"))>1:
			if len(br[0].split("-"))==2:
				if int(br[0].split("-")[1])-int(br[0].split("-")[0])<=1000:
					tt2=int(br[0].split("-")[1])-int(br[0].split("-")[0])
					brA=str(int(br[0].split("-")[0])+tt2/2)
					v1=False
					v2=False
				else:
					brA=br[0]
					v1=True
					v2=False
			if len(br[0].split("-"))==3:
				brA=br[0].split("-",1)[1]
				v1=False
				v2=True
		if len(br[0].split("-"))==1:
			brA=br[0]
			v1=False
			v2=False
		if len(br[1].split("-"))==1:
			brB=br[1]
			v3=False
			v4=False		
		if len(br[1].split("-"))>1:
			if len(br[1].split("-"))==2:
				if int(br[1].split("-")[1])-int(br[1].split("-")[0])<=1000:
					tt2=int(br[1].split("-")[1])-int(br[1].split("-")[0])
					brB=str(int(br[1].split("-")[0])+tt2/2)
					v3=False
					v4=False
				else:
					brB=br[1]
					v3=True
					v4=False
			if len(br[1].split("-"))==3:
				brB=br[1].split("-",1)[1]
				v3=False
				v4=True
		br=[brA,brB]
		if len(br[0].split("-"))>1 or len(br[1].split("-"))>1 or tt=="Unbalanced_translocation":####develop!!!!!!!!!
			b1=read_cytoband(bandd, ch[0], br[0].split("-")[0])#alt
			b2=read_cytoband(bandd, ch[1], br[1].split("-")[0])#alti
			if tt=="Balanced_translocation":
				newname1=["Table 1 - Characterization of the breakpoint region "+cA+b1.split("-")[0]+" of t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=["Table 2 - Characterization of the breakpoint region "+cB+b2.split("-")[0]+" of t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "trans", ch[0])
				name2=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "trans", ch[1])
			elif tt=="Inversion":
				newname1=["Table 1 - Characterization of inv("+cA+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=""
				name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "inv", ch[0])
				name2=name1
			elif tt=="Unbalanced_translocation":
				newname1=["Table 1 - Characterization of the breakpoint region "+cA+b1.split("-")[0]+" of the unbalanced t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=["Table 2 - Characterization of the breakpoint region "+cB+b2.split("-")[0]+" of the unbalanced t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				sz=read_chr_size(chrsz)
				if der==ch[0]:
					name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "trans", ch[0])
					if "p" in b1 and "p" in b2:
						name2=pq_nomenclature(ch[1], ch[1], b1, b2, "1",br[1], False, False, False, False, False, "dup", "")
						br=["1-"+brA,"1-"+brB]
					if "p" in b1 and "q" in b2:
						name2=pq_nomenclature(ch[1], ch[1], b1, b2, br[1], sz[ch[1]], False, False, False, False, False, "dup", "")
						br=["1-"+brA, brB+"-"+sz[ch[1]]]
					if "q" in b1 and "q" in b2:
						name2=pq_nomenclature(ch[1], ch[1], b1, b2, br[1], sz[ch[1]], False, False, False, False, False, "dup", "")
						br=[brA+"-"+sz[ch[0]], brB+"-"+sz[ch[1]]]
					if "q" in b1 and "p" in b2:
						name2=pq_nomenclature(ch[1], ch[1], b1, b2, "1",br[1], False, False, False, False, False, "dup", "")
						br=[brA+"-"+sz[ch[0]], "1-"+brB]
				if der==ch[1]:
					name2=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, v1, v2, v3, v4, "trans", ch[1])
					if "p" in b1 and "p" in b2:
						name1=pq_nomenclature(ch[0], ch[0], b1, b2, "1",br[0], False, False, False, False, False, "dup", "")
						br=["1-"+brA, "1-"+brB]
					if "p" in b1 and "q" in b2:
						name1=pq_nomenclature(ch[0], ch[0], b1, b2, "1",br[0], False, False, False, False, False, "dup", "")
						br=["1-"+brA, brB+"-"+sz[ch[1]]]
					if "q" in b1 and "q" in b2:
						name1=pq_nomenclature(ch[0], ch[0], b1, b2, br[0], sz[ch[0]], False, False, False, False, False, "dup", "")
						br=[brA+"-"+sz[ch[0]], brB+"-"+sz[ch[1]]]
					if "q" in b1 and "p" in b2:
						name1=pq_nomenclature(ch[0], ch[0], b1, b2, br[0], sz[ch[0]], False, False, False, False, False, "dup", "")
						br=[brA+"-"+sz[ch[0]], "1-"+brB]
			deals_with_trans_with_dels(br, tadd, ch, version, newname1, newname2, b1, b2, name1, name2, chrsz)
		else:
			b1=read_cytoband(bandd, ch[0], br[0])#alt
			b2=read_cytoband(bandd, ch[1], br[1])#alti
			if tt=="Balanced_translocation":
				newname1=["Table 1 - Characterization of the breakpoint region "+cA+b1.split("-")[0]+" of t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=["Table 2 - Characterization of the breakpoint region "+cB+b2.split("-")[0]+" of t("+cA+";"+cB+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "trans", ch[0])
				name2=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "trans", ch[1])
			elif tt=="Inversion":
				newname1=["Table 1 - Characterization of inv("+cA+")("+b1.split("-")[0]+";"+b2.split("-")[0]+")","","",""]
				newname2=""
				name1=pq_nomenclature(ch[0], ch[1], b1, b2, br[0], br[1], False, False, False, False, False, "inv", ch[0])
				name2=name1
			deal_with_translocations_inversions(ch, br, version, tadd, newname1, newname2, b1,b2, name1, name2, chrsz)
	return name1, name2

	

def deal_with_translocations_inversions(ch, br, version, tadd, newname1, newname2, b1, b2, name1, name2, chrsz):
	"""deals with simple translocations and inversions. Used by get_results"""
	newt1=parse_tads(select_tads(tadd, ch[0], int(br[0]),chrsz))#alt
	newt2=parse_tads(select_tads(tadd, ch[1], int(br[1]),chrsz))#alt
	intnewt1=ordain(newt1)
	intnewt2=ordain(newt2)
	prepare_exel(newname1, True)
	final_things(intnewt1, ch[0], br[0],version, False, False, b1, name1)#alt
	if intnewt1!=intnewt2:
		if newname2!="":
			make_format_ws2(ws2, False, False)
			prepare_exel(newname2, False)
		final_things(intnewt2, ch[1], br[1], version, False, False, b2, name2)#alt	
	make_format_ws2(ws2, False, False)			

def check_genecard_existence(urll):
	"""Recives a genecard link and confirms if it exists on the site.
		Returns true if exists, returns false if it not exists"""
	bb=['Rscript reqs.r '+urll]
	aa=subprocess.check_output(bb, shell=True, stderr=subprocess.STDOUT)
	if "FALSE" in aa:
		return "true"
	else:
		return "false"		
		
def prepare_exel(newname, first):
	"""makes the header of the new exel"""
	global ws2
	global wb
	if first==True:
		wb=Workbook()
		ws2=wb.active
		ws2.title ="rearrangement A"
	else:
		ws2=wb.create_sheet()
		ws2.title="rearrangement B"
	dat= time.strftime("%d-%m-%Y")
	aa=['TAD-gene content Tool V2.0', '', '', dat, '', '=HYPERLINK("http://192.168.20.16/cgi-bin/tadgctV2.py")']
	ws2.append(aa)
	ws2.append([])
	ws2.append(newname)
	ws2.append([])
	#ws2.append(['Genes GeneCard','Nome do Gene', 'Gene OMIM', 'Fenótipo OMIM',"%HI","Hereditariedade"])
	ws2.append(['Genes', '', '', '', ' ', 'Pathologies', '',''])
	ws2.append(['GeneCard', 'OMIM', 'o/e', '%HI', ' ', 'Associated Phenotype', 'OMIM','Inheritance'])


def read_cytoband(infile, chrr, bp):
	"""Read the cytoband file and the breakpoint,
	and returns the cytoband"""
	f=open(infile)
	dic={}
	for i in f:
		line=i.split()
		if line[0]=="chr"+chrr:#alt
			if int(bp)>int(line[1]) and int(bp)<int(line[2]):
				return line[-1]
	f.close()


def check_points(ws2):
	aa=0
	i=1
	while i<=ws2.max_row:
		n=ws2["A"+str(i)].value
		if n!=None:
			bb=n.encode("UTF-8")
			if bb.startswith("der") or bb.startswith("chr") or bb.startswith("g."):
				aa+=1
		i+=1
	return aa 

def make_background(ws2):
	tt=check_points(ws2)
	i=1
	aa=False
	ist=False
	lets=["A", "B", "C", "D", "E", "F", "G", "H"]
	while i<=ws2.max_row:
		n=ws2["A"+str(i)].value
		if n!=None:
			bb=n.encode("UTF-8")
			if bb.startswith("pter"):
				ist=True
			elif (bb.startswith("der") or bb.startswith("chr") or bb.startswith("g.")) and aa==False and tt!=1:
				aa=True
				if ist==True:
					ws2.delete_rows(i,1)
			elif (bb.startswith("der")==False and bb.startswith("chr")==False and bb.startswith("g.")==False and "TAD" not in bb and "qter" not in bb and "pter" not in bb) and aa==True:
				for el in lets:
					ff=ws2[el+str(i)]
					ff.fill=PatternFill("solid", fgColor="DDDDDD")
			elif (bb.startswith("der") or bb.startswith("chr") or bb.startswith("g.")) and aa==True and tt!=1:
				if ist==False:
					ws2.delete_rows(i,1)
					ist=False
				aa=False
		i+=1
			
def make_format_ws2(ws2, flag, is_region):#flag to show if their is two breakpoints or not
	ws2.append([fntxt])
	"""formats the exel file to a table format"""
	if flag==True and is_region==False:
		make_background(ws2)
	thick_border=Border(bottom=Side(style='thick'))
	large_border=Border(top=Side(style='thick'), bottom=Side(style='thin'))###
	top_border=Border(top=Side(style='thin'),bottom=Side(style='thin'))
	maxborder=Border(top=Side(style='thin'))###
	t_border=Border(bottom=Side(style='thin'), top=Side(style="thick"))#, border_style='dotted')###
	a1=ws2["A1"]
	a1.font=Font(bold=True, size=12, color="808080")
	a1=ws2["D1"]
	a1.font=Font(bold=True, size=12, color="808080")
	a1=ws2["F1"]
	a1.font=Font(bold=True, size=11, color="0000CC")
	a1=ws2['A3']
	a1.font=Font(size=11)
	ws2.merge_cells("A3:H3")
	i=4
	yy=0
	lets=["A", "B", "C", "D", "E", "F", "G", "H"]
	l1=["A", "B", "C", "D", "F", "G", "H"]
	while i<=ws2.max_row:
		n=ws2["E"+str(i)]
		index="A"+str(i)
		aa=ws2[index].value
		if aa!=None:
			bb=aa.encode("UTF-8")
			if "Genes" in bb:
				n=ws2[index]
				n.font=Font(size=12, bold=True)
				n.alignment = Alignment(horizontal='center')
				n=ws2["F"+str(i)]
				n.font=Font(size=12, bold=True)
				n.alignment = Alignment(horizontal='center')
				for el in l1:
					n=ws2[el+str(i-1)]
					rr=ws2[el+str(i+1)]
					rr.border=maxborder
					n.border=thick_border
				n=ws2["E"+str(i)]
				n.border=Border(top=Side(style='thick'))
				ws2.merge_cells("A"+str(i)+":D"+str(i))
				ws2.merge_cells("F"+str(i)+":H"+str(i))
			if ("TAD" in bb and "genome" not in bb) or ("Deleted" in bb and "genome" not in bb) or bb.startswith("qter") or bb.startswith("pter"):
				n=ws2["A"+str(i)]
				n.font=Font(bold=True)
				n2=ws2["B"+str(i)]
				n2.font=Font(bold=True)
				if yy==0:
					for el in lets:
						n=ws2[el+str(i)]
						n.border=large_border
						yy=1
				else:
					for el in lets:
						n=ws2[el+str(i)]
						n.border=t_border					
			elif bb.startswith("GeneC"):
				for wl in lets:
					n=ws2[wl+str(i)]
					n.font=Font(bold=True)
					#n.border=maxborder
				n5=str(i+2)
			elif bb.startswith("der") or bb.startswith("g.") or bb.startswith("chr"):
				n=ws2["A"+str(i)]
				n.font=Font(bold=True)
				ws2.merge_cells("A"+str(i)+":H"+str(i))
			elif "genome" in bb:
				for el in lets:
					n=ws2[el+str(i-1)]
					n.border=thick_border
				n=ws2["A"+str(i)]
				n.font=Font(size=11)
				ws2.merge_cells("A"+str(i)+":H"+str(i))
			else:
				index="B"+str(i)
				aa=ws2[index].value
				if aa!=None:
					bb=aa.encode("UTF-8")
					if bb!="na":
						n=ws2[index]
						n.font=Font(color="0000CC")
                                        else:
                                                n=ws2[index]
    				indexx="G"+str(i)
    				cc=ws2[indexx].value
    				if cc!=None:
						dd=cc.encode("UTF-8")
						if dd!="na":
							n=ws2[indexx]
							n.font=Font(color="0000CC")
				if bb.startswith("TADs are")==False and bb.startswith("The inf")==False:
					n=ws2["A"+str(i)]
					if "HYPERLINK" in n.value:
						n.font=Font(color="0000CC", italic=True)
					else:
						n.font=Font(italic=True)
		elif ws2["G"+str(i)].value!=None:
			n=ws2["G"+str(i)]
			n.font=Font(color="0000CC")
		i+=1
		
def pq_nomenclature(chrA, chrB, bandA, bandB, pqAA, pqBB, isspec, isdelA, isdupA, isdelB, isdupB, typee, der):
	pqA=pqAA.strip()
	pqB=pqBB.strip()
	singA=pqA
	singB=pqB
	if typee=="ins":#aqui considera-se que o chrA corresponde ao cromossoma onde é inserido o DNA, e o chrB o que tem a excisao
		if der==chrA:
			return "g.[chr"+chrA+":"+pqA.replace("-","_")+"inschr"+chrB+":"+pqB.replace("-","_")+"]"
		else:
			return "chr"+chrB+":g."+pqB.replace("-","_")+"del"
	if isspec==True:
		return "chr"+chrA+": g.["+pqAA+"-"+pqBB+"]"
	if isdelA==True:
		aa=pqA.replace("-","_")+"del"
		singA=str(int(pqA.split("-")[1])+1)
	if isdupA==True:
		aa=pqA.replace("-","_")+"dup"
		singA=str(int(pqA.split("-")[1])+1)
	if isdelB==True:
		bb=pqB.replace("-","_")+"del"
		singB=str(int(pqB.split("-")[1])+1)
	if isdupB==True:
		bb=pqB.replace("-","_")+"dup"
		singB=str(int(pqB.split("-")[1])+1)
	if isdupA==False and isdelA==False:
		aa=makePQ(pqA)
	if isdupB==False and isdelB==False and (typee!="del" or typee!="dup"):
		bb=makePQ(pqB)
	if typee=="del":
		return "chr"+chrA+": g."+pqA+"_"+pqB+"del"
	if typee=="dup":
		return "chr"+chrA+": g."+pqA+"_"+pqB+"dup"
	if typee=="inv":
		if "p" in bandA and "p" in bandB:
			if isdelA==True or isdupA==True:
				cc="g.[pter_"+aa+";"+str(int(pqA.split("-")[1])+1)+"_"
			else:
				cc="g.[pter_"+aa+"_"
			if isdelB==True or isdupB==True:
				dd=str(int(pqB.split("-")[0])-1)+"inv;"+bb+"_cen_qter]"
			else:
				dd=bb+"inv"+"_cen_qter]"
		if "q" in bandA and "q" in bandB:
			if isdelA==True or isdupA==True:
				cc="g.[pter_cen_"+aa+";"+str(int(pqA.split("-")[1])+1)+"_"
			else:
				cc="g.[pter_cen_"+aa+"_"
			if isdelB==True or isdupB==True:
				dd=str(int(pqB.split("-")[0])-1)+"inv;"+bb+"_qter]"
			else:
				dd=bb+"inv"+"_qter]"
		if "p" in bandA and "q" in bandB:
			if isdelA==True or isdupA==True:
				cc="g.[pter_"+aa+";"+str(int(pqA.split("-")[1])+1)+"_cen_"
			else:
				cc="g.[pter_"+aa+"_cen_"
			if isdelB==True or isdupB==True:
				dd=str(int(pqB.split("-")[0])-1)+"inv;"+bb+"_qter]"
			else:
				dd=bb+"inv"+"_qter]"
		return cc+dd
	if typee=="trans":
		if der==chrA:
			if "p" in bandA and "p" in bandB:
				cc="der("+chrA+") g.[chr"+chrB+":pter_"+singB+"::chr"+chrA+":"+aa+"_cen_qter]"
			if "p" in bandA and "q" in bandB:
				cc="der("+chrA+") g.[chr"+chrB+":"+singB+"_qterinv::chr"+chrA+":"+aa+"_cen_qter]"
			if "q" in bandA and "p" in bandB:
				cc="der("+chrA+") g.[chr"+chrA+":pter_cen_"+aa+"::chr"+chrB+":pter_"+singB+"inv]"
			if "q" in bandA and "q" in bandB:
				cc="der("+chrA+") g.[chr"+chrA+":pter_cen_"+aa+"::chr"+chrB+":"+singB+"_qter]"
		if der==chrB:
			if "p" in bandA and "p" in bandB:
				cc="der("+chrB+") g.[chr"+chrA+":pter_"+singA+"::chr"+chrB+":"+bb+"_cen_qter]"
			if "p" in bandA and "q" in bandB:
				cc="der("+chrB+") g.[chr"+chrB+":pter_cen"+bb+"::chr"+chrA+":pter_"+singA+"inv]"
			if "q" in bandA and "p" in bandB:
				cc="der("+chrB+") g.[chr"+chrA+":pter_"+singA+"inv::chr"+chrB+":"+bb+"_cen_qter]"
			if "q" in bandA and "q" in bandB:
				cc="der("+chrB+") g.[chr"+chrB+":pter_cen_"+bb+"::chr"+chrA+":"+singA+"_qter]"
		return cc

			
def makePQ(bpk):
	if "-" in bpk:
		gg=bpk.split("-")
		ss=int(gg[1])-int(gg[0])
		return str(int(gg[0])+ss)
	else:
		return bpk

def main(tddss, at1s, tads, tadds, refs, outfiles, cA, cB, brA, brB, version, is_region, tt, der):
	global tdds
	tdds=tddss
	global at1
	at1=at1s
	global tad
	tad=tads
	global tadd#tadfile
	tadd=tadds
	global ref
	ref=refs
	global name 
	name=outfiles
	#texto do final da tabela
	global fntxt
	fntxt="Analysis performed with bioinformatics tool TAD-GCtool v2.0 financiated by the FCT project HMSP-ICT/0016/2013. Human genome version "+version+". TADs according to "+ref+"."
	name1, name2=get_results(cA, cB, brA, brB, version, tadd, is_region, tt, der)
	wb.save(filename=name)
	return name1, name2

if __name__ == "__main__":
    main()

